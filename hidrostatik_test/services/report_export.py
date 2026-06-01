from __future__ import annotations

import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


def export_pdf_report(
    output_path: Path,
    report_title: str,
    metadata: dict[str, Any],
    geometry_data: dict[str, Any],
    test_results: dict[str, Any],
    checklist_data: list[dict[str, Any]],
) -> None:
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        rightMargin=54,
        leftMargin=54,
        topMargin=54,
        bottomMargin=54,
    )
    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#16365D"),
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#243B73"),
        spaceBefore=12,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#333333"),
    )
    bold_body_style = ParagraphStyle(
        "ReportBodyBold",
        parent=body_style,
        fontName="Helvetica-Bold",
    )

    # Title
    story.append(Paragraph(report_title, title_style))
    story.append(Spacer(1, 10))

    # Metadata Section
    story.append(Paragraph("Metadata & Context", section_style))
    meta_table_data = []
    for k, v in metadata.items():
        meta_table_data.append(
            [Paragraph(f"<b>{k}:</b>", body_style), Paragraph(str(v), body_style)]
        )
    t_meta = Table(meta_table_data, colWidths=[150, 350])
    t_meta.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(t_meta)
    story.append(Spacer(1, 15))

    # Geometry Section
    story.append(Paragraph("Pipe Geometry Details", section_style))
    geom_table_data = []
    for k, v in geometry_data.items():
        geom_table_data.append(
            [Paragraph(f"<b>{k}:</b>", body_style), Paragraph(str(v), body_style)]
        )
    t_geom = Table(geom_table_data, colWidths=[150, 350])
    t_geom.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(t_geom)
    story.append(Spacer(1, 15))

    # Test Results Section
    story.append(Paragraph("Hydrostatic Test Evaluation Results", section_style))
    results_table_data = []
    for k, v in test_results.items():
        val_style = bold_body_style if k.lower() in ["passed", "durum", "result"] else body_style
        results_table_data.append(
            [Paragraph(f"<b>{k}:</b>", body_style), Paragraph(str(v), val_style)]
        )
    t_results = Table(results_table_data, colWidths=[150, 350])
    t_results.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(t_results)
    story.append(Spacer(1, 15))

    # Checklist Section
    if checklist_data:
        story.append(Paragraph("Operation Field Checklist", section_style))
        chk_table_data = [[Paragraph("<b>Status</b>", bold_body_style), Paragraph("<b>Control Point / Requirement</b>", bold_body_style)]]
        for item in checklist_data:
            status_text = "[X]" if item.get("checked") else "[ ]"
            chk_table_data.append(
                [
                    Paragraph(status_text, bold_body_style),
                    Paragraph(f"{item.get('label')} ({item.get('ref')})", body_style),
                ]
            )
        t_chk = Table(chk_table_data, colWidths=[60, 440])
        t_chk.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ]
            )
        )
        story.append(t_chk)

    # Build PDF
    doc.build(story)


def export_excel_report(
    output_path: Path,
    report_title: str,
    metadata: dict[str, Any],
    geometry_data: dict[str, Any],
    test_results: dict[str, Any],
    checklist_data: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hydrostatic Test Report"

    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True

    # Styling
    title_font = Font(name="Arial", size=16, bold=True, color="16365D")
    section_font = Font(name="Arial", size=12, bold=True, color="243B73")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True, color="333333")
    regular_font = Font(name="Arial", size=10, color="333333")

    header_fill = PatternFill(start_color="16365D", end_color="16365D", fill_type="solid")
    section_fill = PatternFill(start_color="EAF2FF", end_color="EAF2FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Title
    ws["A1"] = report_title
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30

    current_row = 3

    # Helper function to write simple dicts as table sections
    def write_section(title: str, data: dict[str, Any]) -> None:
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=title).font = section_font
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        current_row += 1

        for k, v in data.items():
            cell_k = ws.cell(row=current_row, column=1, value=k)
            cell_v = ws.cell(row=current_row, column=2, value=str(v))
            cell_k.font = bold_font
            cell_v.font = regular_font
            cell_k.border = thin_border
            cell_v.border = thin_border
            current_row += 1
        current_row += 1  # blank space

    # Write Sections
    write_section("Metadata & Context", metadata)
    write_section("Pipe Geometry Details", geometry_data)
    write_section("Hydrostatic Test Evaluation Results", test_results)

    # Checklist Section
    if checklist_data:
        ws.cell(row=current_row, column=1, value="Operation Field Checklist").font = section_font
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        current_row += 1

        cell_h1 = ws.cell(row=current_row, column=1, value="Status")
        cell_h2 = ws.cell(row=current_row, column=2, value="Control Point / Requirement")
        cell_h1.font = header_font
        cell_h2.font = header_font
        cell_h1.fill = header_fill
        cell_h2.fill = header_fill
        current_row += 1

        for item in checklist_data:
            status_val = "Checked" if item.get("checked") else "Unchecked"
            cell_s = ws.cell(row=current_row, column=1, value=status_val)
            cell_l = ws.cell(row=current_row, column=2, value=f"{item.get('label')} ({item.get('ref')})")
            cell_s.font = bold_font
            cell_l.font = regular_font
            cell_s.border = thin_border
            cell_l.border = thin_border
            current_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # ignore merged cells/titles for width calculation
                if cell.row == 1 or "Checklist" in str(cell.value) or "Metadata" in str(cell.value) or "Results" in str(cell.value) or "Details" in str(cell.value):
                    continue
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(str(output_path))


__all__ = [
    "export_excel_report",
    "export_pdf_report",
]
